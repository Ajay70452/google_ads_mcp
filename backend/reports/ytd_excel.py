"""Build the YTD performance report as an .xlsx workbook."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CURRENT_FILL = PatternFill("solid", fgColor="FFF2CC")
CENTER = Alignment(horizontal="center")

COLUMNS = [
    ("Account", 32, None),
    ("Month", 18, None),
    ("Clicks", 10, "#,##0"),
    ("Impressions", 12, "#,##0"),
    ("CTR", 8, "0.00%"),
    ("Conversions", 12, "#,##0.00"),
    ("Cost", 12, "$#,##0.00"),
    ("Conv. Rate", 11, "0.00%"),
    ("CPL", 11, "$#,##0.00"),
]


def build_ytd_workbook(report: dict) -> bytes:
    """Render the YTD report dict (as returned by /reports/ytd) to xlsx bytes."""
    year = report["year"]
    accounts = report["accounts"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"YTD {year}"

    # Header
    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # Data
    row_idx = 2
    for account in accounts:
        for month in account["months"]:
            values = [
                account["account_name"],
                month["month"],
                month["clicks"],
                month["impressions"],
                # CTR/conv_rate stored as percentages (e.g. 4.32) — divide for Excel %
                (month["ctr"] or 0) / 100,
                month["conversions"],
                month["cost"],
                (month["conv_rate"] or 0) / 100,
                month["cpl"],
            ]
            for col_idx, (value, (_, _, fmt)) in enumerate(zip(values, COLUMNS), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fmt:
                    cell.number_format = fmt
                if month.get("is_current"):
                    cell.fill = CURRENT_FILL
            row_idx += 1

    # Footer note
    note_row = row_idx + 1
    ws.cell(row=note_row, column=1, value="Highlighted rows = current month (data in progress)")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="808080")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
